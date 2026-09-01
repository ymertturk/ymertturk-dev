from fastapi import FastAPI, Depends, Request, Form, UploadFile, File, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from typing import List, Optional
import csv
import io
import re

import models, schemas, crud, db
import uuid
from uuid import getnode as get_mac
from starlette.middleware.sessions import SessionMiddleware
from fastapi.responses import RedirectResponse, Response
from whatsapp_service import WhatsAppService
from telegram_service import TelegramService
from datetime import datetime, timedelta

# Create tables
models.Base.metadata.create_all(bind=db.engine)

app = FastAPI(title="Referee Invitation Tracker")
# Add SessionMiddleware for simple session management
# Secret key is random at startup to force logout on server restart
# max_age=None ensures cookie dies when browser closes
app.add_middleware(SessionMiddleware, secret_key=str(uuid.uuid4()), max_age=None)

import sys
import os

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Mount static files (if we had any, keeping for structure)
static_path = resource_path("static")
if os.path.exists(static_path):
    app.mount("/static", StaticFiles(directory=static_path), name="static")
else:
    # If static folder doesn't exist (e.g. empty and stripped by pyinstaller), make a dummy one or just skip
    # But usually app needs it if templates reference /static.
    # Let's try to create it if we are not frozen (dev), but if frozen we can't write to _MEI usually easily or it won't persist.
    # Better to just skip or warn. If the app needs it for CSS, it will 404 but not crash.
    print(f"Warning: Static directory not found at {static_path}")


templates = Jinja2Templates(directory=resource_path("templates"))

# App Version
APP_VERSION = "v2.0"
templates.env.globals['app_version'] = APP_VERSION

# Start up event to create default users
@app.on_event("startup")
async def startup_event():
    db_session = db.SessionLocal()
    try:
        # Create default admin if not exists
        if not crud.get_user_by_username(db_session, "admin"):
            crud.create_user(db_session, schemas.UserCreate(username="admin", password="admin123", role="admin"))
        
        # Create test user if not exists
        if not crud.get_user_by_username(db_session, "test"):
            crud.create_user(db_session, schemas.UserCreate(username="test", password="test", role="test"))
            
        # Create default user for "New User" scenario if needed, or let Admin create them.
        # Let's create one default user
        if not crud.get_user_by_username(db_session, "selahattinsahin"):
            crud.create_user(db_session, schemas.UserCreate(username="selahattinsahin", password="selahattin1234", role="user"))
            
        # Start Telegram Polling (Background Task)
        import asyncio
        from telegram_service import TelegramService
        
        async def run_polling():
            print("Starting Telegram Polling...")
            temp_db = db.SessionLocal()
            service = TelegramService(temp_db)
            await service.delete_webhook()
            await service.set_my_commands() # Register Menu Button
            temp_db.close()
            
            offset = None
            last_check_time = datetime.min
            
            while True:
                try:
                    poll_db = db.SessionLocal()
                    srv = TelegramService(poll_db)
                    
                    # 1. Process Telegram Updates
                    updates = await srv.get_updates(offset)
                    for update in updates:
                        await srv.process_update(update)
                        offset = update["update_id"] + 1
                    
                    # 2. Automated Checks (Every 1 minute roughly)
                    now = datetime.utcnow()
                    if (now - last_check_time).total_seconds() > 60:
                        last_check_time = now
                        
                        # A. 12 Hour Reminder
                        # Find Pending, sent > 12h ago, reminder_sent=0
                        reminder_time = now - timedelta(hours=12)
                        reminders = poll_db.query(models.Invitation).filter(
                            models.Invitation.status == "PENDING",
                            models.Invitation.sent_at <= reminder_time,
                            models.Invitation.reminder_sent == 0,
                            models.Invitation.referee.has(models.Referee.telegram_chat_id != None)
                        ).all()
                        
                        reminded_referees = []
                        for inv in reminders:
                            try:
                                msg = f"⏳ SAYIN HAKEMİMİZ,\n\nDavetimize cevap vermeniz için son 12 saatiniz kalmıştır. Lütfen 'Kabul Ediyorum' veya 'Mazeretliyim' butonunu kullanınız."
                                await srv.send_message(inv.referee.telegram_chat_id, msg)
                                inv.reminder_sent = 1
                                poll_db.commit()
                                reminded_referees.append(inv.referee.full_name)
                            except Exception as e:
                                print(f"Reminder error inv {inv.id}: {e}")
                        
                        if reminded_referees:
                             admin_config = crud.get_system_config(poll_db, "ADMIN_TELEGRAM_CHAT_ID")
                             if admin_config:
                                 admin_ids = [aid.strip() for aid in admin_config.split(",") if aid.strip()]
                                 r_list = "\n".join([f"- {name}" for name in reminded_referees])
                                 
                                 for aid in admin_ids:
                                     await srv.send_message(aid, f"ℹ️ 12 Saat Hatırlatması Gönderilenler:\n{r_list}")

                        # B. 24 Hour Timeout
                        timeout_time = now - timedelta(hours=24)
                        timeouts = poll_db.query(models.Invitation).filter(
                            models.Invitation.status == "PENDING",
                            models.Invitation.sent_at <= timeout_time,
                            models.Invitation.referee.has(models.Referee.telegram_chat_id != None)
                        ).all()
                        
                        timed_out_referees = []
                        for inv in timeouts:
                            try:
                                inv.status = "TIMED_OUT"
                                inv.notes = (inv.notes or "") + " [SÜRE DOLDU]"
                                poll_db.commit()
                                
                                msg = f"⛔️ SÜRE DOLDU\n\nDavetimize 24 saat içinde cevap vermediğiniz için işlem iptal edilmiştir."
                                await srv.send_message(inv.referee.telegram_chat_id, msg)
                                
                                # Optionally edit original message to say expired
                                if inv.telegram_message_id:
                                    await srv.edit_message_text(inv.referee.telegram_chat_id, inv.telegram_message_id, "⛔️ BU DAVETİN SÜRESİ DOLMUŞTUR.")
                                
                                timed_out_referees.append(inv.referee.full_name)

                            except Exception as e:
                                print(f"Timeout error inv {inv.id}: {e}")
                                
                        if timed_out_referees:
                             admin_config = crud.get_system_config(poll_db, "ADMIN_TELEGRAM_CHAT_ID")
                             if admin_config:
                                 admin_ids = [aid.strip() for aid in admin_config.split(",") if aid.strip()]
                                 t_list = "\n".join([f"- {name}" for name in timed_out_referees])
                                 for aid in admin_ids:
                                     await srv.send_message(aid, f"⛔️ 24 Saat Süresi Dolanlar (İptal Edildi):\n{t_list}")

                    poll_db.close()
                except Exception as e:
                     print(f"Polling loop error: {e}")
                     await asyncio.sleep(5)
                
                await asyncio.sleep(0.5)

        asyncio.create_task(run_polling())

    finally:
        db_session.close()

# Dependency
def get_db():
    db_session = db.SessionLocal()
    try:
        yield db_session
    finally:
        db_session.close()

def get_current_user(request: Request):
    user = request.session.get("user")
    if not user:
        return None
    return user

class NotAuthenticatedException(Exception):
    pass

def login_required(request: Request):
    user = get_current_user(request)
    if not user:
        raise NotAuthenticatedException()
    return user

@app.exception_handler(NotAuthenticatedException)
async def not_authenticated_exception_handler(request: Request, exc: NotAuthenticatedException):
    return RedirectResponse(url="/setup", status_code=303)

# Auth Routes

@app.get("/setup", response_class=HTMLResponse)
def setup_page(request: Request, db: Session = Depends(get_db)):
    # If users exist, redirect to login
    if crud.get_user_by_username(db, "admin"):
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("setup.html", {"request": request})

@app.post("/setup/register")
def setup_register(request: Request, db: Session = Depends(get_db)):
    # Bind current MAC address
    current_mac = str(get_mac())
    crud.set_system_config(db, "bound_mac_address", current_mac)
    return RedirectResponse(url="/login", status_code=303)

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = crud.get_user_by_username(db, username)
    if not user or not crud.verify_password(password, user.hashed_password):
        return templates.TemplateResponse("login.html", {"request": request, "error": "Geçersiz kullanıcı adı veya şifre"})
    
    # Check MAC Address binding for standard users
    if user.role == "user":
        bound_mac = crud.get_system_config(db, "bound_mac_address")
        current_mac = str(get_mac())
        
        if bound_mac:
            if bound_mac != current_mac:
                return templates.TemplateResponse("login.html", {
                    "request": request, 
                    "error": "Lütfen yöneticiyle iletişime geçin. Bu program zaten başka bilgisayarda kayıtlı."
                })
        else:
            # If not bound yet (maybe skipped setup?), bind now?
            # Or should we force setup? Let's bind now to be safe.
            crud.set_system_config(db, "bound_mac_address", current_mac)
            
    # Session setup
    request.session["user"] = {"username": user.username, "role": user.role, "id": user.id}
    return RedirectResponse(url="/", status_code=303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)

@app.get("/admin/settings", response_class=HTMLResponse)
def admin_settings_page(request: Request, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    if user["role"] != "admin":
        return RedirectResponse(url="/", status_code=303)
        
    # Get current config
    config = {
        "TELEGRAM_BOT_TOKEN": crud.get_system_config(db, "TELEGRAM_BOT_TOKEN"),
        "TELEGRAM_BOT_USERNAME": crud.get_system_config(db, "TELEGRAM_BOT_USERNAME"),
        "ADMIN_TELEGRAM_CHAT_ID": crud.get_system_config(db, "ADMIN_TELEGRAM_CHAT_ID")
    }
    
    # Get users for admin management
    users_list = crud.get_users(db)
    
    return templates.TemplateResponse("settings.html", {
        "request": request, 
        "config": config, 
        "user": user,
        "users_list": users_list
    })

@app.post("/admin/settings", response_class=HTMLResponse)
def admin_settings_save(
    request: Request,
    telegram_bot_token: Optional[str] = Form(None),
    telegram_bot_username: Optional[str] = Form(None),
    admin_telegram_chat_id: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    if user["role"] != "admin":
        return RedirectResponse(url="/", status_code=303)
        
    crud.set_system_config(db, "TELEGRAM_BOT_TOKEN", telegram_bot_token.strip() if telegram_bot_token else "")
    crud.set_system_config(db, "TELEGRAM_BOT_USERNAME", telegram_bot_username.strip() if telegram_bot_username else "")
    crud.set_system_config(db, "ADMIN_TELEGRAM_CHAT_ID", admin_telegram_chat_id.strip() if admin_telegram_chat_id else "")
    
    # Reload config and users
    config = {
        "TELEGRAM_BOT_TOKEN": telegram_bot_token,
        "TELEGRAM_BOT_USERNAME": telegram_bot_username,
        "ADMIN_TELEGRAM_CHAT_ID": admin_telegram_chat_id
    }
    users_list = crud.get_users(db)
    
    return templates.TemplateResponse("settings.html", {
        "request": request, 
        "config": config, 
        "user": user,
        "users_list": users_list,
        "message": "Ayarlar başarıyla kaydedildi."
    })

# --- User Management Routes ---

@app.post("/admin/users/add", response_class=RedirectResponse)
def admin_add_user(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    if user["role"] != "admin":
        return RedirectResponse(url="/", status_code=303)
        
    if crud.get_user_by_username(db, username):
        # Error: User exists
        pass # Handle error nicely if we had flash messages, for now just redirect
        
    new_user = schemas.UserCreate(username=username, password=password, role=role)
    crud.create_user(db, new_user)
    return RedirectResponse(url="/admin/settings", status_code=303)

@app.post("/admin/users/{user_id}/delete", response_class=RedirectResponse)
def admin_delete_user(
    request: Request,
    user_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    if user["role"] != "admin":
        return RedirectResponse(url="/", status_code=303)
    
    # Prevent deleting self
    if user["id"] == user_id:
        return RedirectResponse(url="/admin/settings", status_code=303)
        
    crud.delete_user(db, user_id)
    return RedirectResponse(url="/admin/settings", status_code=303)

@app.post("/admin/users/{user_id}/password", response_class=RedirectResponse)
def admin_reset_password(
    request: Request,
    user_id: int,
    new_password: str = Form(...),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    if user["role"] != "admin":
        return RedirectResponse(url="/", status_code=303)
        
    crud.update_user_password(db, user_id, new_password)
    return RedirectResponse(url="/admin/settings", status_code=303)

@app.post("/profile/password", response_class=RedirectResponse)
def change_own_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    # Verify current
    db_user = crud.get_user_by_username(db, user['username'])
    if not db_user or not crud.verify_password(current_password, db_user.hashed_password):
        # Fail
        return RedirectResponse(url="/", status_code=303) # Should ideally show error
        
    crud.update_user_password(db, db_user.id, new_password)
    return RedirectResponse(url="/", status_code=303)

@app.post("/admin/db/import", response_class=HTMLResponse)
async def admin_db_import(
    request: Request,
    db_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    if user["role"] != "admin":
        return RedirectResponse(url="/", status_code=303)

    import shutil
    import os
    from sqlalchemy import create_engine, text
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.automap import automap_base
    
    # 1. Save uploaded file to temp
    temp_filename = f"temp_import_{uuid.uuid4()}.db"
    with open(temp_filename, "wb") as buffer:
        shutil.copyfileobj(db_file.file, buffer)
        
    msg = ""
    try:
        # 2. Connect to Temp DB (SQLite) and Reflect
        temp_db_url = f"sqlite:///{temp_filename}"
        temp_engine = create_engine(temp_db_url)
        
        # Reflect the tables
        TempBase = automap_base()
        TempBase.prepare(autoload_with=temp_engine)
        
        TempSession = sessionmaker(bind=temp_engine)
        temp_session = TempSession()
        
        # Check if 'referees' table exists
        if 'referees' not in TempBase.classes:
             raise Exception("Yüklenen veritabanında 'referees' tablosu bulunamadı.")
             
        OldReferee = TempBase.classes.referees
        # Check for invitation table, might be 'invitations' or something else
        OldInvitation = TempBase.classes.invitations if 'invitations' in TempBase.classes else None
        
        # 3. Read Data and Import
        try:
            old_referees = temp_session.query(OldReferee).all()
            count_new = 0
            count_updated = 0
            
            for old_r in old_referees:
                # Use getattr to safely access fields that exist in old DB
                r_name = getattr(old_r, 'full_name', None) or getattr(old_r, 'name', "Bilinmeyen")
                
                # Normalize
                normalized_name = crud.normalize_name(r_name)
                existing = crud.get_referee_by_normalized_name(db, normalized_name)
                
                # Extract other fields safely
                r_phone = getattr(old_r, 'phone', None)
                r_region = getattr(old_r, 'region', None)
                r_sicil = getattr(old_r, 'sicil_no', None)
                r_cat = getattr(old_r, 'category', None)
                r_tc = getattr(old_r, 'tc_no', None)
                r_iban = getattr(old_r, 'iban', None)
                r_gender = getattr(old_r, 'gender', None)
                r_chat_id = getattr(old_r, 'telegram_chat_id', None)
                
                if existing:
                    # Update basic info if missing in existing
                    existing.phone = r_phone or existing.phone
                    existing.region = r_region or existing.region
                    existing.sicil_no = r_sicil or existing.sicil_no
                    existing.category = r_cat or existing.category
                    existing.tc_no = r_tc or existing.tc_no
                    existing.iban = r_iban or existing.iban
                    existing.gender = r_gender or existing.gender
                    if r_chat_id: # Only update provided chat id
                         existing.telegram_chat_id = r_chat_id
                    
                    count_updated += 1
                    target_referee = existing
                else:
                    # Create
                    new_r = models.Referee(
                        full_name=r_name,
                        full_name_normalized=normalized_name,
                        phone=r_phone,
                        region=r_region,
                        sicil_no=r_sicil,
                        category=r_cat,
                        tc_no=r_tc,
                        iban=r_iban,
                        gender=r_gender,
                        telegram_chat_id=r_chat_id
                    )
                    db.add(new_r)
                    db.flush() # get ID
                    target_referee = new_r
                    count_new += 1
                
                # Import Invitations if table exists
                if OldInvitation and hasattr(old_r, 'invitations_collection'): 
                    # automap usually creates a collection if FK exists. 
                    old_r_id = getattr(old_r, 'id', None)
                    if old_r_id:
                        old_invites = temp_session.query(OldInvitation).filter_by(referee_id=old_r_id).all()
                        for old_inv in old_invites:
                            inv_date = getattr(old_inv, 'event_date', None)
                            inv_name = getattr(old_inv, 'event_name', None)
                            
                            if inv_date: # minimal req
                                exists_inv = db.query(models.Invitation).filter(
                                    models.Invitation.referee_id == target_referee.id,
                                    models.Invitation.event_date == inv_date,
                                    models.Invitation.event_name == inv_name
                                ).first()
                                
                                if not exists_inv:
                                    status_val = getattr(old_inv, 'status', 'PENDING')
                                    new_inv = models.Invitation(
                                        referee_id=target_referee.id,
                                        event_date=inv_date,
                                        event_name=inv_name,
                                        status=status_val,
                                        notes=getattr(old_inv, 'notes', None),
                                        sent_at=getattr(old_inv, 'sent_at', None),
                                        telegram_message_id=getattr(old_inv, 'telegram_message_id', None),
                                        reminder_sent=getattr(old_inv, 'reminder_sent', 0)
                                    )
                                    db.add(new_inv)

             # 4. Import Races directly if table exists
            count_races = 0
            if 'races' in TempBase.classes:
                OldRace = TempBase.classes.races
                try:
                    old_races = temp_session.query(OldRace).all()
                    for old_race in old_races:
                         r_name = getattr(old_race, 'name', None)
                         r_date = getattr(old_race, 'date', None)
                         if r_name and r_date:
                             # Check exist
                             existing_race = db.query(models.Race).filter(models.Race.name == r_name, models.Race.date == r_date).first()
                             if not existing_race:
                                 new_race = models.Race(
                                     name=r_name,
                                     date=r_date,
                                     notes=getattr(old_race, 'notes', None)
                                 )
                                 db.add(new_race)
                                 count_races += 1
                except Exception as e:
                    print(f"Race table import error: {e}")

            # 5. Import System Config (Telegram Settings)
            if 'system_config' in TempBase.classes:
                OldConfig = TempBase.classes.system_config
                try:
                    old_configs = temp_session.query(OldConfig).all()
                    for oc in old_configs:
                        key = getattr(oc, 'key', None)
                        val = getattr(oc, 'value', None)
                        if key and val:
                            if key == "ADMIN_TELEGRAM_CHAT_ID":
                                # Merge with existing
                                current_val = crud.get_system_config(db, key)
                                if current_val:
                                    # Avoid duplicates
                                    current_ids = [x.strip() for x in current_val.split(",") if x.strip()]
                                    new_ids = [x.strip() for x in val.split(",") if x.strip()]
                                    merged = list(set(current_ids + new_ids))
                                    crud.set_system_config(db, key, ",".join(merged))
                                else:
                                    crud.set_system_config(db, key, val)
                            elif key in ["TELEGRAM_BOT_TOKEN", "TELEGRAM_BOT_USERNAME"]:
                                # Overwrite only if not set or user wants restoration primarily
                                # Let's overwrite to ensure full restore
                                crud.set_system_config(db, key, val)
                except Exception as e:
                    print(f"Config import error: {e}")

            db.commit()
            
            # 5. Post-Process: Sync Invitations to Races (Auto-grouping)
            # Find all invitations with null race_id, group by name+date, create race if needed and link.
            orphan_invites = db.query(models.Invitation).filter(models.Invitation.race_id == None).all()
            count_auto_races = 0
            
            for inv in orphan_invites:
                if inv.event_name and inv.event_date:
                    # Find or Create Race
                    race = db.query(models.Race).filter(models.Race.name == inv.event_name, models.Race.date == inv.event_date).first()
                    if not race:
                        race = models.Race(name=inv.event_name, date=inv.event_date)
                        db.add(race)
                        db.flush()
                        count_auto_races += 1
                    
                    inv.race_id = race.id
            
            db.commit()
            
            msg = f"✅ İçe Aktarma Başarılı: {count_new} hakem, {count_updated} güncellendi. {count_races} eski yarış alındı, {count_auto_races} yarış otomatik oluşturuldu."
            
        except Exception as e:
            print(f"Import logic error: {e}")
            msg = f"❌ İçe aktarma sırasında hata oluştu: {e}"
            
        finally:
            temp_session.close()
            
    except Exception as e:
         msg = f"❌ Veritabanı dosyası okunamadı: {e}"
    finally:
        if os.path.exists(temp_filename):
            os.remove(temp_filename)

    # Reload config to display
    config = {
        "TELEGRAM_BOT_TOKEN": crud.get_system_config(db, "TELEGRAM_BOT_TOKEN"),
        "TELEGRAM_BOT_USERNAME": crud.get_system_config(db, "TELEGRAM_BOT_USERNAME"),
        "ADMIN_TELEGRAM_CHAT_ID": crud.get_system_config(db, "ADMIN_TELEGRAM_CHAT_ID")
    }
    
    return templates.TemplateResponse("settings.html", {
        "request": request, 
        "config": config, 
        "user": user, 
        "message": msg
    })



@app.get("/export/excel")
def export_excel(
    request: Request,
    scope_type: str = "all",
    referee_id: Optional[int] = None,
    referee_ids: Optional[List[int]] = Query(None),
    event_filter: Optional[str] = None,
    columns: Optional[List[str]] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    import openpyxl
    from openpyxl.styles import Font
    from io import BytesIO
    from fastapi.responses import Response

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hakemler"

    # Column Mapping
    col_map = {
        "name": ("Ad Soyad", "full_name"),
        "phone": ("Telefon", "phone"),
        "category": ("Kategori", "category"),
        "tc": ("TC No", "tc_no"),
        "iban": ("IBAN", "iban"),
        "gender": ("Cinsiyet", "gender"),
        "region": ("Bölge", "region"),
        "sicil": ("Sicil No", "sicil_no")
    }

    # Determine Headers and Attributes
    headers = []
    attrs = []
    
    if columns:
        for col in columns:
            if col in col_map:
                headers.append(col_map[col][0])
                attrs.append(col_map[col][1])
    else:
        # Default Fallback
        headers = ["Ad Soyad", "Telefon", "Bölge", "Sicil No"]
        attrs = ["full_name", "phone", "region", "sicil_no"]

    # Her halükarda yarış sütunları eklenecek
    headers.extend(["Yarış Adı", "Durum", "Notlar", "Tarih"])

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Fetch data
    query = db.query(models.Referee)
    
    if scope_type == "single" and referee_id:
        query = query.filter(models.Referee.id == referee_id)
    elif scope_type == "selected" and referee_ids:
        query = query.filter(models.Referee.id.in_(referee_ids))
        
    referees = query.order_by(models.Referee.full_name).all()
    
    row_num = 2
    for referee in referees:
        invitations = referee.invitations
        
        # Translation dictionary for statuses
        status_tr = {
            "ACCEPTED": "Kabul",
            "EXCUSED": "Mazeretli",
            "PENDING": "Bekliyor"
        }
        
        # Sadece belirli bir yarış isteniyorsa filtrele
        if event_filter:
            invitations = [inv for inv in invitations if inv.event_name == event_filter]
            
        if not invitations:
            # Yarışı yoksa sadece hakem bilgileriyle tek satır ekle
            for i, attr in enumerate(attrs):
                val = getattr(referee, attr, "")
                ws.cell(row=row_num, column=i+1).value = val
                
            start_col = len(attrs) + 1
            if event_filter:
                ws.cell(row=row_num, column=start_col).value = event_filter
            row_num += 1
        else:
            # Hakemin katıldığı/davet edildiği her bir yarış için ayrı satır oluştur
            for inv in invitations:
                for i, attr in enumerate(attrs):
                    val = getattr(referee, attr, "")
                    ws.cell(row=row_num, column=i+1).value = val
                
                start_col = len(attrs) + 1
                ws.cell(row=row_num, column=start_col).value = inv.event_name
                ws.cell(row=row_num, column=start_col+1).value = status_tr.get(inv.status, inv.status)
                ws.cell(row=row_num, column=start_col+2).value = inv.notes
                ws.cell(row=row_num, column=start_col+3).value = inv.event_date
                row_num += 1

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from urllib.parse import quote
    filename = f"hakemler_{scope_type}.xlsx"
    if event_filter:
        filename = f"{event_filter.replace(' ', '_')}_listesi.xlsx"
        
    encoded_filename = quote(filename)
    headers_http = {
        'Content-Disposition': f"attachment; filename*=utf-8''{encoded_filename}"
    }
    
    return Response(content=buffer.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers_http)

@app.get("/", response_class=HTMLResponse)
def read_root(request: Request, search: Optional[str] = None, event_filter: Optional[str] = None, tab: Optional[str] = "all", page: int = 1, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    limit = 30
    skip = (page - 1) * limit
    
    if tab == "no_invites":
        # For "no_invites", we get all of them (or we needs a specific crud function with pagination)
        # For MVP/Simple app, getting all and filtering or custom query is better.
        # Let's use the crude getAll logic from crud if it was simple list, but we need pagination.
        # We should add a specific query in crud but for now let's do a simple filter in python if dataset small, 
        # OR better, update crud to handle this mode.
        # Let's trust crud.get_referees_without_invitations but that returns ALL. We need to paginate it manually.
        all_no_invites = crud.get_referees_without_invitations(db)
        # Filter by search if needed
        if search:
            search_norm = crud.normalize_name(search)
            all_no_invites = [r for r in all_no_invites if search_norm in r.full_name_normalized]
            
        total_count = len(all_no_invites)
        # Paginate list
        referees = all_no_invites[skip : skip+limit]
    else:
        referees = crud.get_referees(db, search=search, event_filter=event_filter, skip=skip, limit=limit)
        total_count = crud.count_referees(db, search=search, event_filter=event_filter)
        
    total_pages = (total_count + limit - 1) // limit
    
    # Get unique event names for dropdown
    # Result is list of tuples like [('Race A',), ('Race B',)]
    event_names = [e[0] for e in crud.get_unique_event_names(db) if e[0]]
    
    pagination = {
        "page": page,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_prev": page > 1,
        "next_page": page + 1,
        "prev_page": page - 1,
        "total_count": total_count
    }
    
    context = {
        "request": request, 
        "referees": referees, 
        "search": search,
        "event_filter": event_filter,
        "tab": tab,
        "event_names": event_names,
        "pagination": pagination,
        "user": user
    }
    
    if request.headers.get("HX-Request"):
        return templates.TemplateResponse("partials/referee_list.html", context)
    return templates.TemplateResponse("index.html", context)

# ... (create_referee remains same) ...

@app.get("/events/options", response_class=HTMLResponse)
def get_event_options(db: Session = Depends(get_db)):
    event_names = [e[0] for e in crud.get_unique_event_names(db) if e[0]]
    options = ['<option value="">Tüm Yarışlar</option>']
    for event in event_names:
        options.append(f'<option value="{event}">{event}</option>')
    return "\n".join(options)

@app.post("/referees", response_class=HTMLResponse)
def create_referee(
    request: Request,
    full_name: str = Form(...),
    phone: Optional[str] = Form(None),
    region: Optional[str] = Form(None),
    sicil_no: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    try:
        referee_data = schemas.RefereeCreate(full_name=full_name, phone=phone, region=region, sicil_no=sicil_no)

        crud.create_referee(db, referee_data)
        
        # Return first page after creation
        limit = 30
        page = 1
        skip = 0
        
        referees = crud.get_referees(db, skip=skip, limit=limit)
        total_count = crud.count_referees(db)
        total_pages = (total_count + limit - 1) // limit
        
        pagination = {
            "page": page,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1,
            "next_page": page + 1,
            "prev_page": page - 1,
            "total_count": total_count
        }
        
        return templates.TemplateResponse("partials/referee_list.html", {
            "request": request, 
            "referees": referees,
            "pagination": pagination,
            "user": user
        })
    except ValueError as e:
        db.rollback()
        # In a real app we'd return an error message to display
        # For now, just return list to avoid crash, but we need pagination here too
        limit = 30
        referees = crud.get_referees(db, limit=limit)
        total_count = crud.count_referees(db)
        total_pages = (total_count + limit - 1) // limit
        pagination = {"page": 1, "total_pages": total_pages, "has_next": 1 < total_pages, "has_prev": False, "next_page": 2, "prev_page": 0, "total_count": total_count}
        
        return templates.TemplateResponse("partials/referee_list.html", {"request": request, "referees": referees, "pagination": pagination, "user": user})
    except Exception as e:
        db.rollback()
        # Handle duplicate error etc
        limit = 30
        referees = crud.get_referees(db, limit=limit)
        total_count = crud.count_referees(db)
        total_pages = (total_count + limit - 1) // limit
        pagination = {"page": 1, "total_pages": total_pages, "has_next": 1 < total_pages, "has_prev": False, "next_page": 2, "prev_page": 0, "total_count": total_count}
        
        return templates.TemplateResponse("partials/referee_list.html", {"request": request, "referees": referees, "pagination": pagination, "user": user})

@app.get("/referees/{id}", response_class=HTMLResponse)
def read_referee(request: Request, id: int, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    referee = crud.get_referee(db, id)
    if not referee:
        raise HTTPException(status_code=404, detail="Referee not found")
    
    invitations = sorted(referee.invitations, key=lambda x: x.event_date, reverse=True)
    
    stats = {
        "accepted": sum(1 for i in invitations if i.status == "ACCEPTED"),
        "excused": sum(1 for i in invitations if i.status == "EXCUSED"),
        "pending": sum(1 for i in invitations if i.status == "PENDING")
    }
    
    return templates.TemplateResponse("detail.html", {
        "request": request, 
        "referee": referee, 
        "invitations": invitations,
        "stats": stats,
        "user": user
    })


@app.post("/referees/{id}/update", response_class=RedirectResponse)
def update_referee(
    request: Request,
    id: int,
    full_name: str = Form(...),
    phone: Optional[str] = Form(None),
    region: Optional[str] = Form(None),
    sicil_no: Optional[str] = Form(None),
    category: Optional[str] = Form(None),
    tc_no: Optional[str] = Form(None),
    iban: Optional[str] = Form(None),
    gender: Optional[str] = Form(None),
    telegram_chat_id: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    data = {
        "full_name": full_name,
        "phone": phone,
        "region": region,
        "sicil_no": sicil_no,
        "category": category,
        "tc_no": tc_no,
        "iban": iban,
        "gender": gender,
        # Only update telegram_chat_id if provided (admin override?) or keep logic simple
        # "telegram_chat_id": telegram_chat_id 
    }
    # Clean up empty strings to None if needed, or keep as is.
    # For DB updates, it's safer to pass what we get.
    
    crud.update_referee(db, id, data)
    return RedirectResponse(url=f"/referees/{id}", status_code=303)

@app.post("/referees/{id}/delete", response_class=HTMLResponse)
def delete_referee(request: Request, id: int, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    crud.delete_referee(db, id)
    return HTMLResponse("") # Return empty to remove element


@app.post("/referees/{id}/invitations", response_class=HTMLResponse)
def create_invitation(
    request: Request,
    id: int,
    event_date: str = Form(...),
    event_name: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    status: str = Form(...),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):

    # Normalize inputs
    if event_name:
        event_name = event_name.strip()

    # Check for duplicate
    existing_invitation = db.query(models.Invitation).filter(
        models.Invitation.referee_id == id,
        models.Invitation.event_date == event_date,
        models.Invitation.event_name == event_name
    ).first()

    message = None
    if existing_invitation:
        # Duplicate found
        status_map = {"ACCEPTED": "KABUL", "EXCUSED": "MAZERETLİ", "PENDING": "BEKLEMEDE"}
        status_text = status_map.get(existing_invitation.status, existing_invitation.status)
        message = {
            "type": "warning",
            "text": f"⚠️ Bu hakeme bu yarış için zaten davet gönderilmiş: {status_text}"
        }
    else:
        # Create new
        inv_data = schemas.InvitationCreate(
            event_date=event_date,
            event_name=event_name,
            notes=notes,
            status=status
        )
        crud.create_invitation(db, inv_data, id)
        message = {
            "type": "success",
            "text": "✅ Davet başarıyla oluşturuldu."
        }
    
    referee = crud.get_referee(db, id)
    invitations = sorted(referee.invitations, key=lambda x: x.event_date, reverse=True)
    
    # JSON for HX-Trigger
    import json
    trigger_data = {}
    
    if message and message["type"] == "success":
         trigger_data["showMessage"] = {
             "type": "success",
             "title": "Başarılı",
             "text": message["text"]
         }
    elif message and message["type"] == "warning":
         trigger_data["showMessage"] = {
             "type": "warning",
             "title": "Bilgi",
             "text": message["text"]
         }

    headers = {"HX-Trigger": json.dumps(trigger_data)} if trigger_data else {}
    
    return templates.TemplateResponse("partials/invitation_list.html", {
        "request": request, 
        "invitations": invitations,
        "message": message
    }, headers=headers)

@app.post("/invitations/{id}/status", response_class=HTMLResponse)
def update_invitation_status(
    request: Request,
    id: int,
    status: str = Form(...),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    inv = crud.update_invitation_status(db, id, status)
    return templates.TemplateResponse("partials/invitation_row.html", {"request": request, "inv": inv, "user": user})


@app.post("/invitations/{id}/delete", response_class=HTMLResponse)
async def delete_invitation(request: Request, id: int, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    # Check for invitation and retract message if exists
    invitation = crud.get_invitation(db, id)
    if invitation and invitation.telegram_message_id and invitation.referee.telegram_chat_id:
        tg_service = TelegramService(db)
        # Attempt to delete/retract message
        await tg_service.delete_message(invitation.referee.telegram_chat_id, invitation.telegram_message_id)

    crud.delete_invitation(db, id)
    return HTMLResponse("")

@app.post("/referees/message/bulk", response_class=HTMLResponse)
async def bulk_message_referees(
    request: Request,
    referee_ids: list[int] = Form(default=[]),
    message: str = Form(...),
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    if not referee_ids:
        return Response("<div class='alert warning'>Hiç hakem seçilmedi.</div>")
        
    tg_service = TelegramService(db)
    
    # Read file once if exists
    file_bytes = None
    file_name = None
    if file and file.filename:
        file_bytes = await file.read()
        file_name = file.filename

    count_sent = 0
    
    for ref_id in referee_ids:
        referee = crud.get_referee(db, ref_id)
        if not referee or not referee.telegram_chat_id:
            continue
            
        try:
            if file_bytes:
                file_tuple = (file_name, file_bytes, file.content_type)
                await tg_service.send_document(referee.telegram_chat_id, file_tuple, caption=message)
            else:
                await tg_service.send_message(referee.telegram_chat_id, message)
            count_sent += 1
        except Exception as e:
            print(f"Error sending message to {referee.full_name}: {e}")
            
    return Response(f"<div class='alert success'>{count_sent} kişiye mesaj başarıyla gönderildi.</div>")

@app.get("/referees/template/download")
def download_template(request: Request, user: dict = Depends(login_required)):
    import openpyxl
    from openpyxl.styles import Font
    from io import BytesIO
    from fastapi.responses import Response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hakem Sablonu"

    headers = ["Ad Soyad", "Telefon", "Bölge", "Sicil No", "Kategori", "TC No", "IBAN", "Cinsiyet"]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "hakem_yukleme_sablonu.xlsx"
    headers_http = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return Response(content=buffer.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers_http)

@app.post("/referees/import")
async def import_referees(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    import openpyxl
    from io import BytesIO

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        created_count = 0
        updated_count = 0
        
        # Assume Headers in Row 1
        # Map headers to column index
        headers_map = {}
        for cell in ws[1]:
            if cell.value:
                headers_map[cell.value.strip()] = cell.column - 1
                
        # Required header
        if "Ad Soyad" not in headers_map:
             return templates.TemplateResponse("index.html", {
                "request": request, 
                "referees": crud.get_referees(db), 
                 # We need to recreate context basically or just redirect with error param
                 # Redirect is cleaner for now
                "error": "Hata: 'Ad Soyad' sütunu bulunamadı.",
                "user": user,
                 # ... minimal context ...
                 "pagination": {"page":1, "total_pages":1, "total_count":0},
                 "event_names": [] 
            })

        for row in ws.iter_rows(min_row=2, values_only=True):
            full_name = row[headers_map["Ad Soyad"]]
            if not full_name:
                continue
                
            full_name = str(full_name).strip()
            
            # Helper to safely get value
            def get_val(header):
                if header in headers_map and row[headers_map[header]] is not None:
                    return str(row[headers_map[header]]).strip()
                return None

            referee_data = {
                "phone": get_val("Telefon"),
                "region": get_val("Bölge"),
                "sicil_no": get_val("Sicil No"),
                "category": get_val("Kategori"),
                "tc_no": get_val("TC No"),
                "iban": get_val("IBAN"),
                "gender": get_val("Cinsiyet")
            }
            # Remove None values to avoid overwriting existing data with None if cell is empty?
            # Or should empty cell mean delete? 
            # Usually for update import, if empty in excel, maybe keep existing?
            # Let's say if it's explicitly explicitly empty string vs None.
            # openpyxl returns None for empty cells.
            # So, filter out Nones.
            update_data = {k: v for k, v in referee_data.items() if v is not None}
            
            normalized_name = crud.normalize_name(full_name)
            existing_referee = crud.get_referee_by_normalized_name(db, normalized_name)
            
            if existing_referee:
                # Update
                if update_data:
                    crud.update_referee(db, existing_referee.id, update_data)
                    updated_count += 1
            else:
                # Create
                create_data = schemas.RefereeCreate(
                    full_name=full_name,
                    phone=referee_data.get("phone"),
                    region=referee_data.get("region"),
                    sicil_no=referee_data.get("sicil_no"),
                    category=referee_data.get("category"),
                    tc_no=referee_data.get("tc_no"),
                    iban=referee_data.get("iban"),
                    gender=referee_data.get("gender")
                )
                crud.create_referee(db, create_data)
                created_count += 1
                
        return RedirectResponse(url=f"/?success_msg={created_count} Yeni Hakem Eklendi, {updated_count} Hakem Güncellendi", status_code=303)
        
    except Exception as e:
        print(f"Import Error: {e}")
        return RedirectResponse(url="/?error_msg=Dosya işlenirken hata oluştu", status_code=303)


# --- Race Management ---

@app.get("/races", response_class=HTMLResponse)
def list_races(request: Request, page: int = 1, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    limit = 10
    skip = (page - 1) * limit
    races = crud.get_races(db, skip=skip, limit=limit)
    total_races = crud.count_races(db)
    
    # Calculate pages
    total_pages = (total_races + limit - 1) // limit
    if total_pages == 0: total_pages = 1
    
    return templates.TemplateResponse("races.html", {
        "request": request, 
        "races": races, 
        "page": page, 
        "total_pages": total_pages,
        "all_races_for_dropdown": crud.get_races(db, limit=1000), # Fetch all (or many) for the dropdown
        "user": user
    })

@app.post("/races", response_class=HTMLResponse)
def create_race(
    request: Request,
    name: str = Form(...),
    date: str = Form(...),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    try:
        crud.create_race(db, name, date, notes)
        
        # Prepare context for races.html (same as list_races)
        limit = 10
        page = 1
        skip = 0
        races = crud.get_races(db, skip=skip, limit=limit)
        total_races = crud.count_races(db)
        total_pages = (total_races + limit - 1) // limit
        if total_pages == 0: total_pages = 1
        
        return templates.TemplateResponse("races.html", {
            "request": request, 
            "races": races, 
            "page": page, 
            "total_pages": total_pages,
            "all_races_for_dropdown": crud.get_races(db, limit=1000),
            "user": user
        })
    except Exception as e:
        # On error also need to provide context
        limit = 10
        page = 1
        races = crud.get_races(db, skip=0, limit=limit)
        total_races = crud.count_races(db)
        total_pages = (total_races + limit - 1) // limit
        if total_pages == 0: total_pages = 1
        
        return templates.TemplateResponse("races.html", {
            "request": request, 
            "races": races, 
            "error": str(e), 
            "page": page,
            "total_pages": total_pages,
            "all_races_for_dropdown": crud.get_races(db, limit=1000),
            "user": user
        })

@app.post("/races/{id}/update", response_class=HTMLResponse)
def update_race_endpoint(
    request: Request,
    id: int,
    name: str = Form(...),
    date: str = Form(...),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    try:
        crud.update_race(db, id, name, date, notes)
    except Exception as e:
        print(f"Error updating race {id}: {e}")
    
    response = HTMLResponse("")
    response.headers["HX-Redirect"] = "/races"
    return response

@app.post("/races/{id}/delete", response_class=HTMLResponse)
def delete_race(request: Request, id: int, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    crud.delete_race(db, id)
    return HTMLResponse("")

@app.get("/races/{id}", response_class=HTMLResponse)
def read_race(request: Request, id: int, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    race = crud.get_race(db, id)
    if not race:
        return RedirectResponse(url="/races", status_code=303)
        
    # Get all referees to list them for selection
    # Ideally filter out those who already have invitation for THIS race?
    # Or just show them as "Invited".
    referees = crud.get_referees(db, limit=1000) # Get all for now
    
    # Map referee id to invitation status for this race
    race_invites = {inv.referee_id: inv for inv in race.invitations}
    
    # Calculate counts server-side for accuracy
    accepted_count = sum(1 for inv in race.invitations if inv.status == 'ACCEPTED')
    
    return templates.TemplateResponse("race_detail.html", {
        "request": request, 
        "race": race, 
        "referees": referees, 
        "race_invites": race_invites,
        "accepted_count": accepted_count,
        "user": user,
        "now": datetime.utcnow()
    })

@app.post("/races/{id}/bulk-invite", response_class=HTMLResponse)
async def bulk_invite(
    request: Request,
    id: int,
    referee_ids: List[int] = Form(...),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    race = crud.get_race(db, id)
    if not race:
         return Response("Yarış bulunamadı", status_code=404)
         
    tg_service = TelegramService(db)
    results = []
    
    # 1. Create Invitations first
    created_invites = []
    for ref_id in referee_ids:
        # Check if already exists
        existing = db.query(models.Invitation).filter(
            models.Invitation.referee_id == ref_id,
            models.Invitation.race_id == id
        ).first()
        
        if existing:
            created_invites.append(existing)
        else:
            inv_data = schemas.InvitationCreate(
                event_date=race.date,
                event_name=race.name,
                status="PENDING",
                notes=race.notes
            )
            # We need to update create_invitation to handle race_id or do it manually here
            # Update crud.create_invitation first? Or just do it here.
            # Let's do it manually key-val to support race_id
            new_inv = models.Invitation(
                referee_id=ref_id,
                race_id=id,
                event_date=race.date,
                event_name=race.name,
                status="PENDING",
                notes=race.notes
            )
            db.add(new_inv)
            created_invites.append(new_inv)
            
    db.commit()
    
    # 2. Send Telegram Messages
    count_sent = 0
    for inv in created_invites:
        # Refresh to get IDs
        db.refresh(inv)
        
        # Skip if message already sent
        if inv.telegram_message_id:
            continue
            
        if inv.referee.telegram_chat_id:
             res = await tg_service.send_invite(
                 chat_id=inv.referee.telegram_chat_id,
                 referee_name=inv.referee.full_name,
                 event_name=inv.event_name,
                 event_date=inv.event_date,
                 invitation_id=inv.id
             )
             if res["success"]:
                 inv.telegram_message_id = res["message_id"]
                 inv.sent_at = datetime.utcnow()
                 count_sent += 1
                 
    db.commit()
    
    # JSON for HX-Trigger
    import json
    trigger_data = {
        "showMessage": {
            "type": "success",
            "title": "Başarılı",
            "text": f"{count_sent} hakeme davet gönderildi."
        },
        "delayedRefresh": {}
    }
    
    headers = {"HX-Trigger": json.dumps(trigger_data)}
    return Response(f"<div class='alert success'>{count_sent} kişiye davet gönderildi.</div>", headers=headers)

@app.post("/races/{id}/announce", response_class=HTMLResponse)
async def bulk_announce(
    request: Request,
    id: int,
    message: str = Form(...),
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    race = crud.get_race(db, id)
    if not race:
         return Response("Yarış bulunamadı", status_code=404)
         
    tg_service = TelegramService(db)
    
    # Get Accepted Referees for this race
    accepted_invites = [inv for inv in race.invitations if inv.status == "ACCEPTED" and inv.referee.telegram_chat_id]
    
    count_sent = 0
    header = f"📢 DUYURU: {race.name}\n\n"
    final_msg = header + message
    
    # Read file once if exists
    file_bytes = None
    file_name = None
    if file and file.filename:
        file_bytes = await file.read()
        file_name = file.filename

    for inv in accepted_invites:
         try:
             res = None
             if file_bytes:
                 # Send document with caption
                 # Reset implicit cursor or pass bytes directly depending on library support.
                 # httpx inside send_document expects:
                 # files={"document": (filename, bytes, content_type)}
                 file_tuple = (file_name, file_bytes, file.content_type)
                 res = await tg_service.send_document(inv.referee.telegram_chat_id, file_tuple, caption=final_msg)
             else:
                 # Text only
                 res = await tg_service.send_message(inv.referee.telegram_chat_id, final_msg)

             if res["success"]:
                 count_sent += 1
         except Exception as e:
             print(f"Announcement Error for {inv.referee.full_name}: {e}")
             
    # Return success alert
    return Response(f"<div class='alert success'>{count_sent} hakeme duyuru gönderildi.</div>")

@app.get("/import", response_class=HTMLResponse)
def import_page(request: Request, user: dict = Depends(login_required)):
    return templates.TemplateResponse("import.html", {"request": request, "user": user})

def process_import_row(db: Session, full_name, phone, region, sicil_no=None, event_name=None, status=None, notes=None):
    if not full_name or len(full_name.strip()) < 2:
        return False, "Geçersiz isim"
    
    # Try to find existing referee
    referee = None
    
    # 1. By Sicil No
    if sicil_no:
        referee = db.query(models.Referee).filter(models.Referee.sicil_no == sicil_no).first()
        
    # 2. By Name if not found yet
    if not referee:
        norm_name = crud.normalize_name(full_name)
        referee = crud.get_referee_by_normalized_name(db, norm_name)
        
    if referee:
        # UPDATE existing
        changed = False
        if phone and phone != referee.phone:
            referee.phone = phone
            changed = True
        if region and region != referee.region:
            referee.region = region
            changed = True
        if sicil_no and sicil_no != referee.sicil_no:
            referee.sicil_no = sicil_no
            changed = True
            
        if changed:
            db.commit()
            crud.backup_database()
            db.refresh(referee)
            
        action_result = "Güncellendi"
    else:
        # CREATE new
        try:
            referee = crud.create_referee(db, schemas.RefereeCreate(full_name=full_name, phone=phone, region=region, sicil_no=sicil_no))
            action_result = "Eklendi"
        except Exception as e:
            db.rollback()
            return False, str(e)

    # Handle Invitation if event provided
    if event_name and referee:
        # Check if invitation exists
        invitation = None
        for inv in referee.invitations:
            if inv.event_name == event_name:
                invitation = inv
                break
        
        target_status = status.upper() if status else "PENDING"
        # Map common status text to Enum if needed
        if "KABUL" in target_status and "EDEM" not in target_status and "RED" not in target_status:
            target_status = "ACCEPTED"
        elif "MAZERET" in target_status or "RED" in target_status or "HAYIR" in target_status:
            target_status = "EXCUSED"
        elif "PENDING" not in target_status and "ACCEPTED" not in target_status and "EXCUSED" not in target_status:
             # Keep as is if valid or fallback
             pass

        if invitation:
            # Update invitation
            if invitation.status != target_status or (notes and invitation.notes != notes):
                invitation.status = target_status
                if notes: invitation.notes = notes
                db.commit()
                crud.backup_database()
        else:
            # Create invitation
            # Need event_date? If importing from export, we should have it in file or we use today/dummy?
            # If exported, event_date column exists.
            # But process_import_row args doesn't have it yet.
            # Let's default to today if missing or use existing logic.
            # Ideally we extract event_date from file too.
            from datetime import date
            crud.create_invitation(db, schemas.InvitationCreate(event_name=event_name, event_date=date.today(), status=target_status, notes=notes), referee.id)

    return True, action_result


@app.post("/import/paste", response_class=HTMLResponse)
def import_paste(request: Request, paste_data: str = Form(...), db: Session = Depends(get_db), user: dict = Depends(login_required)):
    lines = paste_data.split('\n')
    result = {"added": 0, "skipped": 0, "errors": 0, "error_details": []}
    
    for line in lines:
        line = line.strip()
        if not line: continue
        
        # Try different delimiters
        if '\t' in line: parts = line.split('\t')
        elif ';' in line: parts = line.split(';')
        elif ',' in line: parts = line.split(',')
        else: parts = [line]
        
        parts = [p.strip() for p in parts]
        full_name = parts[0]
        phone = parts[1] if len(parts) > 1 else None
        region = parts[2] if len(parts) > 2 else None
        
        # Paste usually doesn't have sicil_no, event info etc.
        success, error = process_import_row(db, full_name, phone, region)
        if success:
            if error == "Güncellendi":
                 # Maybe count as added or create new category "updated"?
                 result["added"] += 1
            else:
                 result["added"] += 1
        elif error == "Mükerrer (Zaten var)":
             # Should not happen with Upsert logic anymore unless strictly create mode
             result["skipped"] += 1
        else:
            result["errors"] += 1
            result["error_details"].append(f"{line}: {error}")
            
    return templates.TemplateResponse("import.html", {"request": request, "result": result, "user": user})

@app.post("/import/csv", response_class=HTMLResponse)
async def import_csv(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), user: dict = Depends(login_required)):
    content = await file.read()
    text = content.decode('utf-8')
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    
    # Normalize headers
    if reader.fieldnames:
        reader.fieldnames = [name.lower().strip() for name in reader.fieldnames]
    
    result = {"added": 0, "skipped": 0, "errors": 0, "error_details": []}
    
    for row in reader:
        # Map common variations
        full_name = row.get('full_name') or row.get('name') or row.get('fullname') or row.get('ad soyad')
        phone = row.get('phone') or row.get('mobile') or row.get('tel') or row.get('telefon')
        region = row.get('region') or row.get('district') or row.get('area') or row.get('bölge') or row.get('il')
        sicil_no = row.get('sicil_no') or row.get('sicil') or row.get('no')
        event_name = row.get('event_name') or row.get('event') or row.get('yaris')
        status = row.get('status') or row.get('durum')
        notes = row.get('notes') or row.get('notlar') or row.get('açıklama')
        
        if not full_name:
            result["errors"] += 1
            continue
            
        success, error = process_import_row(db, full_name, phone, region, sicil_no, event_name, status, notes)
        if success:
            result["added"] += 1
        else:
            result["errors"] += 1
            result["error_details"].append(f"{full_name}: {error}")

    return templates.TemplateResponse("import.html", {"request": request, "result": result, "user": user})

@app.post("/import/drive", response_class=HTMLResponse)
async def import_drive(request: Request, drive_url: str = Form(...), db: Session = Depends(get_db), user: dict = Depends(login_required)):
    import httpx
    
    # Transform URL if it's a standard viewer URL
    # https://docs.google.com/spreadsheets/d/KEY/edit#gid=0 -> https://docs.google.com/spreadsheets/d/KEY/export?format=csv
    url = drive_url
    if "docs.google.com/spreadsheets" in url:
        match = re.search(r"/d/([a-zA-Z0-9-_]+)", url)
        if match:
            doc_id = match.group(1)
            url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv"
            
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url, follow_redirects=True)
            response.raise_for_status()
            content = response.text
            
        f = io.StringIO(content)
        reader = csv.DictReader(f)
        
        # Normalize headers
        if reader.fieldnames:
            reader.fieldnames = [name.lower().strip() for name in reader.fieldnames]
        
        result = {"added": 0, "skipped": 0, "errors": 0, "error_details": []}
        
        for row in reader:
             # Map common variations (same as CSV import)
            full_name = row.get('full_name') or row.get('name') or row.get('fullname') or row.get('ad soyad')
            phone = row.get('phone') or row.get('mobile') or row.get('tel') or row.get('telefon')
            region = row.get('region') or row.get('district') or row.get('area') or row.get('bölge') or row.get('il')
            sicil_no = row.get('sicil_no') or row.get('sicil') or row.get('no')
            event_name = row.get('event_name') or row.get('event') or row.get('yaris')
            status = row.get('status') or row.get('durum')
            notes = row.get('notes') or row.get('notlar') or row.get('açıklama')
            
            if not full_name:
                result["errors"] += 1
                continue
                
            success, error = process_import_row(db, full_name, phone, region, sicil_no, event_name, status, notes)
            if success:
                result["added"] += 1
            else:
                result["errors"] += 1
                result["error_details"].append(f"{full_name}: {error}")

        return templates.TemplateResponse("import.html", {"request": request, "result": result, "user": user})
        
    except Exception as e:
        return templates.TemplateResponse("import.html", {"request": request, "result": {"added": 0, "skipped": 0, "errors": 1, "error_details": [f"Hata: {str(e)}"]}, "user": user})



@app.get("/invitations/bulk-update", response_class=HTMLResponse)
def bulk_update_page(request: Request, user: dict = Depends(login_required)):
    return templates.TemplateResponse("bulk_status.html", {"request": request, "user": user})

@app.post("/invitations/bulk-update", response_class=HTMLResponse)
async def process_bulk_update(
    request: Request, 
    event_date: str = Form(...), 
    event_name: Optional[str] = Form(None),
    paste_data: Optional[str] = Form(None), 
    drive_url: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    lines = []
    if drive_url:
        import httpx
        url = drive_url
        if "docs.google.com/spreadsheets" in url:
            match = re.search(r"/d/([a-zA-Z0-9-_]+)", url)
            if match:
                doc_id = match.group(1)
                url = f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv"
        
        try:
            async with httpx.AsyncClient() as client:
                response = await client.get(url, follow_redirects=True)
                response.raise_for_status()
                content = response.text
                lines = content.split('\n')
        except Exception as e:
             return templates.TemplateResponse("bulk_status.html", {"request": request, "error": f"Google Drive Hatası: {str(e)}", "user": user})

    elif paste_data:
        lines = paste_data.split('\n')
    
    if not lines:
         return templates.TemplateResponse("bulk_status.html", {"request": request, "error": "Lütfen veri yapıştırın veya Google Drive linki girin.", "user": user})

    result = {"updated": 0, "not_found": 0, "no_invitation": 0, "details": []}
    
    target_date = None
    try:
        from datetime import datetime
        target_date = datetime.strptime(event_date, "%Y-%m-%d").date()
    except:
        return templates.TemplateResponse("bulk_status.html", {"request": request, "error": "Geçersiz tarih formatı.", "user": user})

    for line in lines:
        line = line.strip()
        if not line: continue
        
        # Split by tab or common separators. Google Sheets uses tab.
        if '\t' in line: parts = line.split('\t')
        elif ',' in line: parts = line.split(',') # Fallback
        else: parts = [line]
        
        if len(parts) < 2:
            continue # Skip lines without status
            
        name_part = parts[0].strip()
        status_part = parts[-1].strip().upper() # Take last part as status
        
        # Normalize name to find referee
        norm_name = crud.normalize_name(name_part)
        referee = crud.get_referee_by_normalized_name(db, norm_name)
        
        if not referee:
            result["not_found"] += 1
            result["details"].append(f"Bulunamadı: {name_part}")
            continue
            
        # Find invitation for this date
        invitation = None
        for inv in referee.invitations:
            if inv.event_date == target_date:
                invitation = inv
                break
        
        # Determine status
        new_status = "PENDING"
        # Check for rejection first since "Kabul edemiyorum" contains "Kabul"
        if "MAZERET" in status_part or "HAYIR" in status_part or "RED" in status_part or "KABUL EDE" in status_part:
            new_status = "EXCUSED"
        elif "KABUL" in status_part:
            new_status = "ACCEPTED"
            
        if invitation:
            # Update event name if provided
            if event_name:
                invitation.event_name = event_name
                
            if new_status != "PENDING":
                crud.update_invitation_status(db, invitation.id, new_status)
                result["updated"] += 1
            else:
                # Even if status didn't change, we might have updated event_name
                if event_name:
                    db.commit()
                    result["updated"] += 1
                else:
                    result["details"].append(f"Durum Değişmedi: {name_part}")
        else:
            # Create new invitation if it doesn't exist
            inv_data = schemas.InvitationCreate(
                event_date=target_date,
                event_name=event_name if event_name else "Google Form Import",
                status=new_status
            )
            crud.create_invitation(db, inv_data, referee.id)
            result["updated"] += 1 # Count as updated/handled

    return templates.TemplateResponse(request=request, name="bulk_status.html", context={"result": result, "user": user})

@app.get("/admin/reset", response_class=HTMLResponse)
def reset_page(request: Request, user: dict = Depends(login_required)):
    if user['role'] != 'admin' and user['username'] != 'selahattinsahin':
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse("reset.html", {"request": request, "user": user})

@app.post("/admin/reset", response_class=HTMLResponse)
def perform_reset(request: Request, confirmation: str = Form(...), db: Session = Depends(get_db), user: dict = Depends(login_required)):
    if user['role'] != 'admin' and user['username'] != 'selahattinsahin':
        return RedirectResponse(url="/", status_code=303)
        
    if confirmation != "HEPSİNİ SİL":
        return templates.TemplateResponse("reset.html", {"request": request, "error": "Hatalı onay ifadesi. Lütfen 'HEPSİNİ SİL' yazın.", "user": user})
    
    crud.delete_all_data(db)
    return RedirectResponse(url="/", status_code=303)


# --- WhatsApp Integration ---

@app.post("/invitations/{id}/whatsapp")
def send_whatsapp_invite(
    request: Request,
    id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    invitation = db.query(models.Invitation).filter(models.Invitation.id == id).first()
    if not invitation:
        raise HTTPException(status_code=404, detail="Davet bulunamadı")
        
    referee = invitation.referee
    if not referee.phone:
        return templates.TemplateResponse("partials/invitation_row.html", {"request": request, "inv": invitation, "error": "Telefon numarası yok"})

    wa_service = WhatsAppService(db)
    result = wa_service.send_invite(
        to_number=referee.phone,
        referee_name=referee.full_name,
        event_name=invitation.event_name,
        event_date=invitation.event_date
    )
    
    if result["success"]:
        invitation.whatsapp_message_sid = result["sid"]
        invitation.sent_at = datetime.utcnow()
        db.commit()
        # Optionally mark status as emailed/contacted or just leave pending
        # Let's keep PENDING but maybe add a note?
    else:
        # Show error?
        pass
        
    return templates.TemplateResponse("partials/invitation_row.html", {"request": request, "inv": invitation, "wa_result": result})

@app.post("/webhook/whatsapp")
async def whatsapp_webhook(request: Request, db: Session = Depends(get_db)):
    form_data = await request.form()
    from_number = form_data.get("From", "").replace("whatsapp:", "")
    body = form_data.get("Body", "").strip().upper()
    
    if not from_number:
        return Response(content="No from number", media_type="text/plain")

    # Clean from_number (Twilio sends +905...)
    # Our DB might have 05... or +90...
    # Simple match: check if DB phone ends with the last 10 digits of incoming
    short_number = from_number[-10:] # last 10 digits
    
    # Inefficient but works for small DB. For larger DB, store normalized phones.
    referees = db.query(models.Referee).all()
    target_referee = None
    for ref in referees:
        if ref.phone and ref.phone.replace(" ", "").replace("-", "")[-10:] == short_number:
            target_referee = ref
            break
            
    if not target_referee:
        return Response(content="<Response><Message>Sizi sistemde bulamadım.</Message></Response>", media_type="application/xml")
        
    # Find latest PENDING invitation for this referee
    # Or find invitation sent roughly recently? 
    # Let's just take the latest pending one.
    invitation = db.query(models.Invitation).filter(
        models.Invitation.referee_id == target_referee.id,
        models.Invitation.status == "PENDING"
    ).order_by(models.Invitation.created_at.desc()).first()
    
    if not invitation:
         # Check if they already replied?
         return Response(content="<Response><Message>Bekleyen bir davetiniz bulunmamaktadır.</Message></Response>", media_type="application/xml")

    # Update logic
    if "KABUL" in body or "EVET" in body:
        new_status = "ACCEPTED"
        reply_msg = "Teşekkürler, kaydınız alındı."
    elif "MAZERET" in body or "HAYIR" in body or "RED" in body:
        new_status = "EXCUSED"
        reply_msg = "Bilgilendirme için teşekkürler."
    else:
        return Response(content="<Response><Message>Lütfen 'KABUL' veya 'MAZERET' yazınız.</Message></Response>", media_type="application/xml")
        
    invitation.status = new_status
    invitation.last_response_at = datetime.utcnow()
    invitation.response_content = body
    db.commit()
    crud.backup_database()
    
    return Response(content=f"<Response><Message>{reply_msg}</Message></Response>", media_type="application/xml")

# --- Telegram Integration ---

@app.post("/invitations/{id}/telegram")
async def send_telegram_invite(
    request: Request,
    id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(login_required)
):
    invitation = db.query(models.Invitation).filter(models.Invitation.id == id).first()
    if not invitation:
        raise HTTPException(status_code=404, detail="Davet bulunamadı")
        
    referee = invitation.referee
    if not referee.telegram_chat_id:
        return templates.TemplateResponse("partials/invitation_row.html", {"request": request, "inv": invitation, "error": "Telegram bağlı değil"})

    tg_service = TelegramService(db)
    tg_service = TelegramService(db)
    result = await tg_service.send_invite(
        chat_id=referee.telegram_chat_id,
        referee_name=referee.full_name,
        event_name=invitation.event_name,
        event_date=invitation.event_date,
        invitation_id=invitation.id
    )
    
    if result["success"]:
        invitation.telegram_message_id = result["message_id"]
        invitation.sent_at = datetime.utcnow()
        db.commit()
    
    return templates.TemplateResponse("partials/invitation_row.html", {"request": request, "inv": invitation, "tg_result": result})

@app.post("/webhook/telegram")
def telegram_cleanup():
    return {"ok": True, "info": "Webhooks are disabled. System uses polling now."}


@app.post("/whatsapp/timeout-check", response_class=HTMLResponse)
def check_whatsapp_timeouts(request: Request, db: Session = Depends(get_db), user: dict = Depends(login_required)):
    # Find invitations sent > 24 hours ago that are still PENDING
    limit_time = datetime.utcnow() - timedelta(hours=24)
    
    overdue_invites = db.query(models.Invitation).filter(
        models.Invitation.status == "PENDING",
        models.Invitation.sent_at != None,
        models.Invitation.sent_at < limit_time
    ).all()
    
    count = 0
    for inv in overdue_invites:
        # Mark as 'TIMED_OUT' or just notify? User asked to mark/notify.
        # Let's add a special status or just assume Pending (Overdue) in UI.
        # For this request: "24 saat... cevap vermeyenler Beklemede... yöneticiye bilgi verilecek"
        # Since they are already Pending, we just need to list them for the admin.
        count += 1
        
    return f"<span>{count} adet zaman aşımı kontrol edildi.</span>"


if __name__ == "__main__":
    import uvicorn
    import webbrowser
    from threading import Timer

    def open_browser():
        webbrowser.open("http://127.0.0.1:8000")

    Timer(1.5, open_browser).start()
    # reload=False is required for frozen apps
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)
