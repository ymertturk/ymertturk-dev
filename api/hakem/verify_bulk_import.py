import crud
from db import SessionLocal
import models
import schemas
import openpyxl
import os
import openpyxl
import os

TEMPALTE_FILE = "test_import.xlsx"

def create_test_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    
    headers = ["Ad Soyad", "Telefon", "Bölge", "Sicil No", "Kategori", "TC No", "IBAN", "Cinsiyet"]
    ws.append(headers)
    
    # 1 New, 1 Existing (Update)
    ws.append(["BULK TEST USER 1", "5551112233", "Ankara", "999888", "Ulusal", "11111111111", "TR123", "Kadın"])
    ws.append(["ABDULLAH TURAN", "5559998877", "İzmir", "UPD-SICIL", "Uluslararası", "22222222222", "TR456", "Erkek"])
    
    wb.save(TEMPALTE_FILE)
    print(f"Created test excel: {TEMPALTE_FILE}")

def verify_import():
    db = SessionLocal()
    try:
        # Create Excel
        create_test_excel()
        
        # Simulate "Import" logic
        print("Running Import Logic Simulation...")
        wb = openpyxl.load_workbook(TEMPALTE_FILE)
        ws = wb.active
        
        # Map headers
        headers_map = {}
        for cell in ws[1]:
            if cell.value:
                headers_map[cell.value.strip()] = cell.column - 1

        created_count = 0
        updated_count = 0
        
        # iter_rows(min_row=2)
        for row in ws.iter_rows(min_row=2, values_only=True):
            full_name = row[headers_map["Ad Soyad"]]
            if not full_name:
                continue
                
            full_name = str(full_name).strip()
            
             # Helper to safely get value
            def get_val(header):
                if header in headers_map and row[headers_map[header]] is not None:
                    return str(row[headers_map[header]]).strip()
                return None

            referee_data = {
                "phone": get_val("Telefon"),
                "region": get_val("Bölge"),
                "sicil_no": get_val("Sicil No"),
                "category": get_val("Kategori"),
                "tc_no": get_val("TC No"),
                "iban": get_val("IBAN"),
                "gender": get_val("Cinsiyet")
            }
            
            # Filter None (simulating API logic)
            update_data = {k: v for k, v in referee_data.items() if v is not None}

            normalized_name = crud.normalize_name(full_name)
            existing_referee = crud.get_referee_by_normalized_name(db, normalized_name)
            
            if existing_referee:
                print(f"Updating existing: {full_name}")
                crud.update_referee(db, existing_referee.id, update_data)
                updated_count += 1
            else:
                print(f"Creating new: {full_name}")
                create_data = schemas.RefereeCreate(
                    full_name=full_name,
                    phone=referee_data.get("phone"),
                    region=referee_data.get("region"),
                    sicil_no=referee_data.get("sicil_no"),
                    category=referee_data.get("category"),
                    tc_no=referee_data.get("tc_no"),
                    iban=referee_data.get("iban"),
                    gender=referee_data.get("gender")
                )
                crud.create_referee(db, create_data)
                created_count += 1
        
        print(f"Done. Created: {created_count}, Updated: {updated_count}")
        
        # Verify New User
        new_user = crud.get_referee_by_normalized_name(db, crud.normalize_name("BULK TEST USER 1"))
        if new_user and new_user.phone == "5551112233":
             print("SUCCESS: New user created correctly.")
        else:
             print("FAIL: New user not found or incorrect.")

        # Verify Updated User
        updated_user = crud.get_referee_by_normalized_name(db, crud.normalize_name("ABDULLAH TURAN"))
        if updated_user and updated_user.phone == "5559998877" and updated_user.sicil_no == "UPD-SICIL":
             print("SUCCESS: Existing user updated correctly.")
        else:
             print("FAIL: Existing user not updated correctly.")
             print(f"Got: {updated_user.phone}, {updated_user.sicil_no}")

    except Exception as e:
        print(f"VERIFICATION FAILED: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()
        if os.path.exists(TEMPALTE_FILE):
             os.remove(TEMPALTE_FILE)

if __name__ == "__main__":
    verify_import()
